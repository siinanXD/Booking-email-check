"""Excel-Export (.xlsx) für den Putzplan.

Layout laut docs/WHATSAPP_BOT_SPEC.md: dunkelblauer Header mit weißer
Schrift, AutoFilter, fixierte Kopfzeile, Zebra-Streifen, rote Warnzeile
bei Same-Day-Wechsel (Check-out und nächster Check-in am selben Tag)
und Fußzeile mit Erstellungszeitpunkt.
"""

from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.features.cleaning.models import (
    CleaningPartner,
    CleaningTask,
    CleaningTaskStatus,
)

_STATUS_LABELS: dict[CleaningTaskStatus, str] = {
    CleaningTaskStatus.UNASSIGNED: "Offen (kein Partner)",
    CleaningTaskStatus.SCHEDULED: "Geplant",
    CleaningTaskStatus.NOTIFIED: "Benachrichtigt",
    CleaningTaskStatus.DONE: "Erledigt",
    CleaningTaskStatus.CANCELLED: "Storniert",
}

_HEADERS = [
    "Datum",
    "Wochentag",
    "Objekt",
    "Zimmer",
    "Gast",
    "Check-out",
    "Nächster Check-in",
    "Zeitfenster",
    "Zugewiesen an",
    "Telefon",
    "Status",
    "Bemerkung",
    "Erledigt",
]

_WEEKDAYS = [
    "Montag",
    "Dienstag",
    "Mittwoch",
    "Donnerstag",
    "Freitag",
    "Samstag",
    "Sonntag",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
_WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
_WARN_FONT = Font(color="9C0006")
# Stornierte Zeilen bleiben sichtbar (grau + durchgestrichen) statt zu
# verschwinden — sonst liest der Kunde "da fehlt eine Buchung", wo in
# Wahrheit ein Storno vorliegt.
_CANCELLED_FONT = Font(color="808080", strike=True)


def status_label(status: CleaningTaskStatus) -> str:
    """Deutsches Label für einen Auftragsstatus."""
    return _STATUS_LABELS.get(status, status.value)


def _fmt_date(value: date | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


def _weekday(value: date | None) -> str:
    return _WEEKDAYS[value.weekday()] if value else ""


def _next_checkins(tasks: list[CleaningTask]) -> dict[str, date | None]:
    """Je Auftrag: frühester Check-in derselben Wohnung nach dem Putztermin."""
    result: dict[str, date | None] = {}
    for task in tasks:
        candidates = [
            other.check_in
            for other in tasks
            if other.task_id != task.task_id
            # Ein stornierter Gast reist nicht an — er darf weder ein
            # Zeitfenster verengen noch eine Same-Day-Warnung auslösen.
            and other.status != CleaningTaskStatus.CANCELLED
            and other.property_name == task.property_name
            and other.check_in is not None
            and task.cleaning_date is not None
            and other.check_in >= task.cleaning_date
        ]
        result[task.task_id] = min(candidates) if candidates else None
    return result


def _window_label(cleaning_date: date | None, next_checkin: date | None) -> str:
    if cleaning_date is None or next_checkin is None:
        return ""
    days = (next_checkin - cleaning_date).days
    if days <= 0:
        return "\u26a0\ufe0f gleicher Tag"
    return f"{days} Tag{'e' if days != 1 else ''}"


def build_cleaning_xlsx(
    tasks: list[CleaningTask],
    partners_by_id: dict[str, CleaningPartner],
    *,
    now: datetime | None = None,
) -> bytes:
    """Erzeugt eine formatierte Putzplan-Arbeitsmappe als Bytes."""
    wb = Workbook()
    ws = wb.active
    weeks = {t.cleaning_date.isocalendar()[1] for t in tasks if t.cleaning_date}
    ws.title = f"Putzplan KW{weeks.pop()}" if len(weeks) == 1 else "Putzplan"

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    next_checkin = _next_checkins(tasks)
    for row_idx, task in enumerate(tasks, start=2):
        cancelled = task.status == CleaningTaskStatus.CANCELLED
        # Storno: Zeitfenster, Partner und Checkbox bleiben leer \u2014 die Zeile
        # dokumentiert die entfallene Reinigung, sie beauftragt niemanden.
        partner = None if cancelled else partners_by_id.get(task.partner_id or "")
        checkin_next = None if cancelled else next_checkin.get(task.task_id)
        window = _window_label(task.cleaning_date, checkin_next)
        if task.status == CleaningTaskStatus.DONE:
            checkbox = "\u2611"
        else:
            checkbox = "" if cancelled else "\u2610"
        ws.append(
            [
                _fmt_date(task.cleaning_date),
                _weekday(task.cleaning_date),
                task.property_name or "",
                task.room_number or "",
                task.guest_name or "",
                _fmt_date(task.check_out),
                _fmt_date(checkin_next),
                window,
                partner.name if partner else "",
                partner.phone if partner else "",
                status_label(task.status),
                task.note or "",
                checkbox,
            ]
        )
        same_day = window.startswith("\u26a0")
        for cell in ws[row_idx]:
            if cancelled:
                cell.font = _CANCELLED_FONT
            elif same_day:
                cell.fill = _WARN_FILL
                cell.font = _WARN_FONT
            elif row_idx % 2 == 0:
                cell.fill = _ZEBRA_FILL

    last_col = get_column_letter(len(_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(tasks) + 1, 2)}"
    _autosize(ws, len(_HEADERS))

    stamp = (now or datetime.now(UTC)).strftime("%d.%m.%Y %H:%M")
    footer_row = len(tasks) + 3
    footer = ws.cell(row=footer_row, column=1)
    footer.value = f"Erstellt am {stamp} via Booking-Email Assistent"
    footer.font = Font(italic=True, color="808080")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _ics_escape(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\n", "\\n")
    )


def build_cleaning_ics(
    tasks: list[CleaningTask],
    partners_by_id: dict[str, CleaningPartner],
    *,
    now: datetime,
) -> bytes:
    """Erzeugt einen iCal-Kalender (.ics) mit je einem Ganztags-Termin pro Auftrag."""
    stamp = now.strftime("%Y%m%dT%H%M%SZ")
    lines: list[str] = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Booking-Email//Putzplan//DE",
        "CALSCALE:GREGORIAN",
    ]
    for task in tasks:
        if task.status == CleaningTaskStatus.CANCELLED or task.cleaning_date is None:
            continue
        day = task.cleaning_date.strftime("%Y%m%d")
        next_day = (task.cleaning_date + timedelta(days=1)).strftime("%Y%m%d")
        partner = partners_by_id.get(task.partner_id or "")
        summary = f"Putzen: {task.property_name or '—'}"
        desc = f"Gast: {task.guest_name or '—'} | Status: {status_label(task.status)}"
        if partner:
            desc += f" | Partner: {partner.name}"
        if task.note:
            desc += f" | Bemerkung: {task.note}"
        lines += [
            "BEGIN:VEVENT",
            f"UID:cleaning-{task.task_id}@booking-email",
            f"DTSTAMP:{stamp}",
            f"DTSTART;VALUE=DATE:{day}",
            f"DTEND;VALUE=DATE:{next_day}",
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(desc)}",
            "END:VEVENT",
        ]
    lines.append("END:VCALENDAR")
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def _autosize(ws: object, columns: int) -> None:
    """Setzt grobe Spaltenbreiten anhand des Inhalts."""
    for col_idx in range(1, columns + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in ws[letter]:  # type: ignore[index]
            value = cell.value
            if value is not None:
                longest = max(longest, len(str(value)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 40)  # type: ignore[attr-defined]
