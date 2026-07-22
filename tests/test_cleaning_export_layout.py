"""Tests: Excel-Layout des Putzplans (Spec-Design)."""

from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO

from openpyxl import load_workbook

from backend.features.cleaning.export import build_cleaning_xlsx
from backend.features.cleaning.models import (
    CleaningPartner,
    CleaningTask,
    CleaningTaskStatus,
)

_NOW = datetime(2026, 7, 8, 12, 0, tzinfo=UTC)


def _task(
    task_id: str,
    *,
    cleaning_date: date,
    check_in: date | None = None,
    check_out: date | None = None,
    status: CleaningTaskStatus = CleaningTaskStatus.SCHEDULED,
    prop: str = "FeWo Seeblick",
) -> CleaningTask:
    return CleaningTask(
        task_id=task_id,
        property_name=prop,
        guest_name="Max",
        check_in=check_in,
        check_out=check_out or cleaning_date,
        cleaning_date=cleaning_date,
        partner_id="p1",
        status=status,
    )


def _partners() -> dict[str, CleaningPartner]:
    return {
        "p1": CleaningPartner(partner_id="p1", name="Petra", phone="+4915722222222")
    }


def _load(tasks: list[CleaningTask]):
    data = build_cleaning_xlsx(tasks, _partners(), now=_NOW)
    return load_workbook(BytesIO(data)).active


def test_sheet_title_mit_kalenderwoche() -> None:
    ws = _load([_task("t1", cleaning_date=date(2026, 7, 15))])
    assert ws.title == "Putzplan KW29"


def test_header_design_und_freeze() -> None:
    ws = _load([_task("t1", cleaning_date=date(2026, 7, 15))])
    header = ws[1][0]
    assert header.value == "Datum"
    assert header.font.bold is True
    assert header.font.color.rgb.endswith("FFFFFF")
    assert header.fill.fgColor.rgb.endswith("1F4E79")
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_zeile_mit_wochentag_partner_und_checkbox() -> None:
    ws = _load(
        [
            _task(
                "t1",
                cleaning_date=date(2026, 7, 15),  # Mittwoch
                status=CleaningTaskStatus.DONE,
            )
        ]
    )
    row = [c.value for c in ws[2]]
    assert row[0] == "15.07.2026"
    assert row[1] == "Mittwoch"
    assert row[2] == "FeWo Seeblick"
    assert row[8] == "Petra"
    assert row[12] == "\u2611"  # erledigt


def test_naechster_checkin_und_same_day_warnung() -> None:
    """Folgt der nächste Gast am Putztag, wird die Zeile rot markiert."""
    tasks = [
        _task(
            "t1",
            cleaning_date=date(2026, 7, 15),
            check_out=date(2026, 7, 15),
        ),
        _task(
            "t2",
            cleaning_date=date(2026, 7, 20),
            check_in=date(2026, 7, 15),  # Same-Day-Anreise nach t1
            check_out=date(2026, 7, 20),
        ),
    ]
    ws = _load(tasks)
    row = [c.value for c in ws[2]]
    assert row[6] == "15.07.2026"  # nächster Check-in
    assert "gleicher Tag" in row[7]
    assert ws[2][0].fill.fgColor.rgb.endswith("FFC7CE")


def test_zeitfenster_in_tagen() -> None:
    tasks = [
        _task("t1", cleaning_date=date(2026, 7, 15), check_out=date(2026, 7, 15)),
        _task(
            "t2",
            cleaning_date=date(2026, 7, 25),
            check_in=date(2026, 7, 18),
            check_out=date(2026, 7, 25),
        ),
    ]
    ws = _load(tasks)
    assert [c.value for c in ws[2]][7] == "3 Tage"


def test_stornierte_zeile_durchgestrichen_und_ohne_auftrag() -> None:
    """Storno bleibt sichtbar dokumentiert, beauftragt aber niemanden."""
    ws = _load(
        [
            _task(
                "t1",
                cleaning_date=date(2026, 7, 15),
                status=CleaningTaskStatus.CANCELLED,
            )
        ]
    )
    row = [c.value for c in ws[2]]
    assert row[2] == "FeWo Seeblick"
    assert row[10] == "Storniert"
    assert row[8] is None or row[8] == ""  # kein Partner beauftragt
    assert row[12] is None or row[12] == ""  # keine Erledigt-Checkbox
    assert ws[2][0].font.strike is True
    assert ws[2][0].font.color.rgb.endswith("808080")


def test_stornierter_gast_loest_keine_same_day_warnung_aus() -> None:
    """Ein stornierter Anreise-Gast verengt kein Zeitfenster."""
    tasks = [
        _task(
            "t1",
            cleaning_date=date(2026, 7, 15),
            check_out=date(2026, 7, 15),
        ),
        _task(
            "t2",
            cleaning_date=date(2026, 7, 20),
            check_in=date(2026, 7, 15),  # Same-Day, aber storniert
            check_out=date(2026, 7, 20),
            status=CleaningTaskStatus.CANCELLED,
        ),
    ]
    ws = _load(tasks)
    row = [c.value for c in ws[2]]
    assert row[6] in (None, "")  # kein "nächster Check-in" vom Storno-Gast
    assert row[7] in (None, "")
    assert not ws[2][0].fill.fgColor.rgb.endswith("FFC7CE")


def test_fusszeile_mit_erstellungszeit() -> None:
    ws = _load([_task("t1", cleaning_date=date(2026, 7, 15))])
    footer = ws.cell(row=4, column=1).value
    assert footer == "Erstellt am 08.07.2026 12:00 via Booking-Email Assistent"
